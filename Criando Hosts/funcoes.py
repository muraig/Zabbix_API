import openpyxl
from estrutura_API_ZABBIX import *
import json
import requests
import pandas as pd

class Le_Tabela():
    def __init__(self, tabela):
        doc_planilha = openpyxl.load_workbook(tabela)
        self.planilha = doc_planilha['Hosts']
        self.hosts = []
        self.le_tabela()

    def le_tabela(self):
        index = ['Host', 'Visible_Name', 'Grupos', 'Interfaces_Tipo', 'Interfaces_IP', 'Interfaces_Porta',
                 'Descricao', 'Templates']
        dados_hosts =[]
        for num_linha in range(4, self.planilha.max_row + 1):
            validacao_linha = True
            for celula in range(2, self.planilha.max_column + 1):
                verificacao_celula = self.verifica_celula(num_linha, celula)
                if not verificacao_celula:
                    validacao_linha = False
                    break
                else:
                    validacao_linha = True
            if validacao_linha == True:
                dados_hosts.append(self.le_linha(num_linha))
                self.hosts = pd.DataFrame(data=dados_hosts, columns=index)

    def le_linha(self, num_linha):
        host = self.planilha.cell(num_linha, 2).value
        visible_name = self.planilha.cell(num_linha, 3).value
        grupos = (self.planilha.cell(num_linha, 4).value).split(',')
        interfaces_tipo = (self.planilha.cell(num_linha, 5).value).split(',')
        interfaces_ip = (self.planilha.cell(num_linha, 6).value).split(',')
        interfaces_porta = str((self.planilha.cell(num_linha, 7).value)).split(',')
        descricao = self.planilha.cell(num_linha, 8).value
        templates = (self.planilha.cell(num_linha, 9).value).split(',')
        dados_host = [host, visible_name, grupos, interfaces_tipo, interfaces_ip, interfaces_porta, descricao, templates]
        return dados_host

    def verifica_celula(self, numero_linha, numero_coluna):
        valor_celula = self.planilha.cell(numero_linha, numero_coluna).value
        if not valor_celula:
            print('Verifique a linha {} célula {}'.format(numero_linha, numero_coluna))
            return None
        else:
            return valor_celula

class Zabbix():
    header = {'content-type': 'application/json'}

    def __init__(self, usuario, senha, endereco_url):
        self.url = endereco_url
        autenticacao["params"]["user"] = usuario
        autenticacao["params"]["password"] = senha

        resposta_zbx = self.envia_comando_json(autenticacao)

        if "error" in resposta_zbx:
            result = resposta_zbx["error"]["data"]
            print(result)
            print("Verifique url, header, user ou password !!!")
            print("Código sendo finalizado")
            exit()
        else:
            self.token = resposta_zbx["result"]
            # result = "Token gerado com sucesso"

    def envia_comando_json(self, comando):
        res = requests.post(self.url, data=json.dumps(comando), headers=self.header)
        resposta = json.dumps(res.json(), indent=4, sort_keys=True)
        resultado = json.loads(resposta)
        return resultado

    def consulta_id_grupo_hosts(self, nome_Grupo):
        consultas["hostgroup"]["params"]["filter"]["name"] = nome_Grupo
        consultas["hostgroup"]["auth"] = self.token
        resposta_zbx = self.envia_comando_json(consultas["hostgroup"])
        if not resposta_zbx['result']:
            result = None
        else:
            result = resposta_zbx['result'][0]['groupid']
        return result

    def cria_grupo(self, nome_grupo):
        id_grupo = self.consulta_id_grupo_hosts(nome_grupo)
        if id_grupo == None:
            criacao["hostgroup"]["params"]["name"] = nome_grupo
            criacao["hostgroup"]["auth"] = self.token
            resposta_zbx = self.envia_comando_json(criacao["hostgroup"])
            if "error" in resposta_zbx:
                print('Erro cria Host Group', nome_grupo)
                result = resposta_zbx['error']['data']
            else:
                result = resposta_zbx['result']['groupids'][0]
            return result
        return id_grupo

    def consulta_id_hosts(self, host_name):
        consultas['hostname']['params']['filter']['name'] = host_name
        consultas["hostname"]["auth"] = self.token
        resposta_zbx = self.envia_comando_json(consultas['hostname'])
        if not resposta_zbx['result']:
            result = None
        else:
            result = resposta_zbx['result'][0]['hostid']
        return result


    def cria_host(self, hosts_DataFrame):
        for index, linha in hosts_DataFrame.hosts.iterrows():
            interfaces = []
            grupos = []
            id_grupo = {'groupid': None}

            for numero_interface in range(0, len(linha['Interfaces_Tipo'])):
                tipo = linha['Interfaces_Tipo'][numero_interface].replace(' ', '')
                ip = linha['Interfaces_IP'][numero_interface].replace(' ', '')
                porta = linha['Interfaces_Porta'][numero_interface].replace(' ', '')

                interface_dict['ip'] = ip
                interface_dict['port'] = porta
                if numero_interface == 0:
                    interface_dict['main'] = 1
                else:
                    interface_dict['main'] = 1

                if tipo == 'Agente':
                    interface_dict['type'] = 1
                    try:
                        del interface_dict['details']
                    except:
                        pass
                elif tipo == 'SNMP':
                    interface_dict['type'] = 2
                    interface_dict['details'] = {
                        'version': 2,
                        'bulk': 1,

                        'community': '{$SNMP_COMMUNITY}'}

                interfaces.append(interface_dict.copy())

            for grupo in linha['Grupos']:
                id_grupo['groupid'] = self.cria_grupo(grupo)
                grupos.append(id_grupo.copy())

            criacao['host']['params']['host'] = linha['Host']
            criacao['host']['params']['name'] = linha['Visible_Name']
            criacao['host']['params']['interfaces'] = interfaces
            criacao['host']['params']['groups'] = grupos
            criacao['host']['auth'] = self.token
            print(criacao['host'])
            resposta_zbx = self.envia_comando_json(criacao['host'])
            print(resposta_zbx)
